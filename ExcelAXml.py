
# TODO : El programa automatiza la tarea de realizar un XML con algunos datos especificos de un archivo de excel.
# TODO : DESARROLLADO POR SEBASTIAN EZEQUIEL EPSTEIN.
# TODO : Ultima modificación 17/08/2022

# CUANDO PIDA EL EXCEL: Instancias SAP Diciembre 2021 (SAPlogon v1).xlsx
# CUANDO PIDA DONDE GUARDAR EL XML: C:/Users/sebie/AppData/Roaming/SAP/Common/

import openpyxl
import xml.etree.cElementTree as ET
import uuid as _uuid
from datetime import datetime
import os

def crearXML(excel_a_leer, ruta_XML):
    
    # VARIABLES
    paisElegido = ''
    empresaElegida = ''
    _baja = ''
    sid = ''
    numero = ''
    _server = ''
    _serviceid = ''
    _routerid = ''
    _router = ''
    empresaDescripcion = ''
    _SAProuter = ''
    _MessageServer = ''
    _groupServer = ''
    _msid = ''
    _hostName = ''
    _familia = ''
    i = 0
    fechaUpdated = datetime.today().strftime('%Y-%m-%d %H:%M')
    fechaBackUp = datetime.today().strftime('%d-%m-%Y--%H.%M')


    def crearItem():
        if(empresaElegida != 'None' and paisElegido != 'None' and empresaElegida != 'PRIMAX'
        and sid != 'None' and sid != '-' and sid != ''
        and numero != 'None' and numero != '-' and numero != ''
        and _baja == 'None'
        and _familia != 'WEBDISP' and _familia != '#N/D' and _familia != 'BW'
        and _familia != 'PI' and _familia != 'PO' and _familia != 'TREX'
        and _familia != 'PORTAL' and _familia != 'NW' and _familia != 'HANA'):
            return True
        else:
            return False


    def crearConMessage():
        if(_MessageServer != 'None' and _groupServer != 'None'
        and paisElegido != 'None' and empresaElegida != 'None'):
            return True
        else:
            return False


    def crearService():
        if(sid != 'None' and sid != '-' and sid != ''
        and numero != 'None' and numero != '-' and numero != ''
        and paisElegido != 'None' and empresaElegida != 'None'
        and _baja == 'None'
        and _familia != 'WEBDISP' and _familia != '#N/D' and _familia != 'BW'
        and _familia != 'PI' and _familia != 'PO' and _familia != 'TREX'
        and _familia != 'PORTAL' and _familia != 'NW' and _familia != 'HANA'):
            return True
        else:
            return False


    def crearComentario():
        if(_comentario != 'None' and _comentario != '-'):
            Memo = ET.SubElement(Service, 'Memo',
                                attrib={'xml:space': 'preserve'})
            Memo.text = _comentario


    def crearRouter():
        Router = ET.SubElement(Routers, 'Router',
                            uuid=_routerid,
                            name=_router,
                            description=_router,
                            router=_router)


    def crearMessageserver():
        Messageserver = ET.SubElement(Messageservers, "Messageserver",
                                    uuid=_msid,
                                    name=sid,
                                    description=sid,
                                    host=_MessageServer,
                                    port='36' + numero
                                    )


    def findfile(name, path):        
        for dirpath, dirname,filename in os.walk(path):
            if name in filename:
                # os.remove("backup.xml")
                os.rename("SAPUILandscape.xml", f"backup-{fechaBackUp}.xml")
                return os.path.join(dirpath, name)

    # CUAL EXCEL LEER
    # excelALeer = str(
    #     input('Por favor ingrese el excel (con el .xlsx) que quiere convertir a XML: '))
    book = openpyxl.load_workbook(
        # excelALeer
        # 'Instancias SAP Diciembre 2021 (SAPlogon v2 control cruzado).xlsx', data_only=True)
        excel_a_leer, data_only=True)
    hoja = book.active


    # DONDE SE GUARDA EL XML
    # para abrirlo directo en SAPLOGON
    # ruta = 'C:/Users/sebie/AppData/Roaming/SAP/Common/'
    # Abrirlo en la carpeta del proyecto
    # ruta = 'C:/Users/sebie/Desktop/Proyecto111/'

    Landscape = ET.Element("Landscape", updated=fechaUpdated, version='1',
                        generator="SAP GUI for Windows v7700.1.6.156")


    # CREAR ARBOL DE XML
    Workspaces = ET.SubElement(Landscape, "Workspaces")
    Workspace = ET.SubElement(Workspaces, "Workspace",
                            uuid=str(_uuid.uuid4()), name="Local", expanded="1")
    Services = ET.SubElement(Landscape, "Services")
    Routers = ET.SubElement(Landscape, "Routers")
    Messageservers = ET.SubElement(Landscape, "Messageservers")

    # RECORRER EL EXCEL
    for fila in range(2, hoja.max_row+1):
        for columna in range(1, hoja.max_column+1):
            cell_obj = hoja.cell(row=fila, column=columna)
            if(columna == 2):  # PAIS
                if(paisElegido != str(cell_obj.value) and str(cell_obj.value) != 'None'):
                    Node = ET.SubElement(Workspace, 'Node', uuid=str(
                        _uuid.uuid4()), name=str(cell_obj.value), expanded='1')
                    paisElegido = str(cell_obj.value)
            if(columna == 4): #DADO DE BAJA
                _baja = str(cell_obj.value)            
            if(columna == 5):  # EMPRESA
                if(empresaElegida != str(cell_obj.value) and str(cell_obj.value) != 'None' and str(cell_obj.value) != 'PRIMAX'):
                    Node1 = ET.SubElement(Node, 'Node', uuid=str(
                        _uuid.uuid4()), name=str(cell_obj.value))
                    empresaElegida = str(cell_obj.value)
                empresaElegida = str(cell_obj.value)
            if(columna == 7):
                _hostName = str(cell_obj.value)
            if(columna == 9):  # SID: Systemid
                sid = str(cell_obj.value)
            if(columna == 10):  # NUMERO
                numero = str(cell_obj.value)
            if(columna == 11):  # DESCRIPCION
                empresaDescripcion = _hostName + ' / ' + str(cell_obj.value)
            if(columna == 13):  # FAMILIA
                _familia = str(cell_obj.value).upper()
            if(columna == 22):  # IP
                _server = str(cell_obj.value) + ':32' + numero
            if(columna == 23):  # SAP ROUTER
                _SAProuter = str(cell_obj.value)
            if(columna == 25):  # MessageServer
                _MessageServer = str(cell_obj.value)
            if(columna == 26):  # Group Server
                _groupServer = str(cell_obj.value)
            if(columna == 27):  # Comentario
                _comentario = str(cell_obj.value)
                if(crearItem()):
                    _serviceid = str(_uuid.uuid4())
                    Item = ET.SubElement(Node1, 'Item', uuid=str(
                        _uuid.uuid4()), serviceid=_serviceid)
                if(_SAProuter == 'None'):
                    if(crearConMessage()):
                        # SI TIENEN AMBOS, SE CREA UN SERVICE CON UN ID PARA COMUNICARLO CON EL MessageServer
                        _msid = str(_uuid.uuid4())
                        if(crearService()):
                            Service = ET.SubElement(Services, 'Service',
                                                    type='SAPGUI',
                                                    uuid=_serviceid,
                                                    name=empresaDescripcion,
                                                    systemid=sid,
                                                    msid=_msid,
                                                    mode='1',
                                                    server=_server,
                                                    sncop="-1",
                                                    sapcpg="1100",
                                                    dcpg="2")
                            crearMessageserver()
                            crearComentario()
                    else:
                        # SI NO TIENEN AMBOS, SE CREA UN SERVICE SIN NADA PARA EL MessageServer
                        if(crearService()):
                            Service = ET.SubElement(Services, 'Service',
                                                    type='SAPGUI',
                                                    uuid=_serviceid,
                                                    name=empresaDescripcion,
                                                    systemid=sid,
                                                    mode='1',
                                                    server=_server,
                                                    sncop="-1",
                                                    sapcpg="1100",
                                                    dcpg="2")
                            crearComentario()
                else:  # if(str(cell_obj.value) != 'None'):
                    _router = '/H/' + _SAProuter
                    _routerid = str(_uuid.uuid4())
                    if(crearConMessage()):
                        # SI TIENEN AMBOS, SE CREA UN SERVICE CON UN ID PARA COMUNICARLO CON EL MessageServer
                        _msid = str(_uuid.uuid4())
                        if(crearService()):
                            Service = ET.SubElement(Services, 'Service',
                                                    type='SAPGUI',
                                                    uuid=_serviceid,
                                                    name=empresaDescripcion,
                                                    systemid=sid,
                                                    msid=_msid,
                                                    mode='1',
                                                    server=_server,
                                                    routerid=_routerid,
                                                    sncop="-1",
                                                    sapcpg="1100",
                                                    dcpg="2")
                            crearRouter()
                            crearMessageserver()
                            crearComentario()
                    else:
                        if(crearService()):
                            Service = ET.SubElement(Services, 'Service',
                                                    type='SAPGUI',
                                                    uuid=_serviceid,
                                                    name=empresaDescripcion,
                                                    systemid=sid,
                                                    mode='1',
                                                    server=_server,
                                                    routerid=_routerid,
                                                    sncop="-1",
                                                    sapcpg="1100",
                                                    dcpg="2")
                            crearRouter()
                            crearComentario()


    print('-----------------')
    print('Fila:', fila)
    print('Columna:', columna)
    print('Ejecutado, XML CREADO!')

    findfile("SAPUILandscape.xml", ruta_XML)
    archivo = ET.ElementTree(Landscape)
    archivo.write(ruta_XML + "SAPUILandscape.xml")
    
    

crearXML('PruebaExcel (v1).xlsx','C:/Users/sebie/Desktop/Proyecto1/')